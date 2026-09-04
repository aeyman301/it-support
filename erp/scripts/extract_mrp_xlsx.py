#!/usr/bin/env python3
"""
Extracts materials, outstanding purchase orders, and production plan demand
from PNA Technologies' monthly "MRP PLAN ALL" workbook into a single JSON
file that scripts/import_to_firestore.mjs can load.

Usage:
    python3 extract_mrp_xlsx.py <path-to-xlsx> [--limit N] [--out out.json]

The workbook's layout (sheet names, header rows) is a recurring monthly
report format from the supply chain team, but the MRP sheet's *column
positions* shift from month to month (extra columns get inserted, e.g. a
"shortage" or "Incoming 1st week" column). Column indices are therefore
detected dynamically from the header row each run rather than hardcoded:
the incoming-PO and demand-forecast columns are found by scanning for
date-valued header cells, split into two groups by the "Plan to order"
label column, with the very first date column (the "STOCK AS <date>"
snapshot) explicitly excluded since it's a stock snapshot, not a schedule
entry.
"""

import argparse
import json
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

RAW_STOCK_SHEET = "1. Raw Stock - Warehouse"
RAW_STOCK_HEADER_ROW = 6
RAW_STOCK_FIRST_DATA_ROW = 7
COL_CODE = 2  # COMPONENT PART NO
COL_DESC = 4  # DESCRIPTION
COL_SUPPLIER = 5  # SUPPLIER
COL_UOM = 7  # UOM (B0M)
COL_ONHAND = 12  # STOCK TAKE CLOSING

MRP_SHEET = "MRP"
MRP_HEADER_ROW = 9
MRP_FIRST_DATA_ROW = 10
MRP_COL_CODE = 2  # PNA PART NO
PLAN_TO_ORDER_LABEL = "plan to order"

DEFAULT_MIN_ORDER_QTY = 1


def sanitize_id(value) -> str:
    text = str(value).strip()
    text = text.replace("/", "-")
    return text or "UNKNOWN"


def month_end_iso(dt) -> str | None:
    if not isinstance(dt, (datetime, date)):
        return None
    return dt.strftime("%Y-%m-%d")


def as_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    return None


def find_mrp_date_columns(ws, header_row: int, max_col: int):
    """Returns (incoming_cols, demand_cols) as {col_index: iso_date} dicts,
    detected from the header row rather than hardcoded positions. The first
    date column found is the "STOCK AS <date>" snapshot and is excluded;
    date columns before the "Plan to order" label are incoming-PO schedule,
    date columns after it are the demand forecast.
    """
    header = [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]

    plan_to_order_col = None
    for i, val in enumerate(header, start=1):
        if isinstance(val, str) and val.strip().lower() == PLAN_TO_ORDER_LABEL:
            plan_to_order_col = i
            break
    if plan_to_order_col is None:
        raise ValueError(f'Could not find a "{PLAN_TO_ORDER_LABEL}" column in the MRP header row')

    first_date_col = None
    incoming_cols = {}
    demand_cols = {}
    for i, val in enumerate(header, start=1):
        iso = month_end_iso(val)
        if not iso:
            continue
        if first_date_col is None:
            first_date_col = i  # the "STOCK AS <date>" snapshot — excluded
            continue
        if i < plan_to_order_col:
            incoming_cols[i] = iso
        elif i > plan_to_order_col:
            demand_cols[i] = iso

    return incoming_cols, demand_cols


def extract(path: str, limit: int | None):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Materials (+ on-hand stock + supplier) ----
    ws = wb[RAW_STOCK_SHEET]
    materials = []
    seen_codes = set()
    for r in range(RAW_STOCK_FIRST_DATA_ROW, ws.max_row + 1):
        raw_code = ws.cell(row=r, column=COL_CODE).value
        if raw_code in (None, ""):
            continue
        code = sanitize_id(raw_code)
        if code in seen_codes:
            continue  # duplicate part number row, keep the first
        seen_codes.add(code)

        desc = ws.cell(row=r, column=COL_DESC).value or code
        supplier = ws.cell(row=r, column=COL_SUPPLIER).value
        supplier = str(supplier).strip() if supplier not in (None, 0, "0") else ""
        uom = ws.cell(row=r, column=COL_UOM).value or "pcs"
        onhand = as_number(ws.cell(row=r, column=COL_ONHAND).value) or 0

        material = {
            "id": code,
            "code": code,
            "name": str(desc).strip(),
            "uom": str(uom).strip(),
            "leadTimeDays": 0,
            "safetyStock": 0,
            "minOrderQty": DEFAULT_MIN_ORDER_QTY,
            "onHandQty": round(onhand),
        }
        if supplier:
            material["supplier"] = supplier
        materials.append(material)
        if limit and len(materials) >= limit:
            break

    material_ids = {m["id"] for m in materials}

    # ---- Purchase orders + production plan, from the MRP sheet ----
    ws2 = wb[MRP_SHEET]
    incoming_dates, demand_dates = find_mrp_date_columns(ws2, MRP_HEADER_ROW, ws2.max_column)

    today_iso = date.today().isoformat()
    purchase_orders = []
    production_plan = []
    matched = 0
    skipped = 0

    for r in range(MRP_FIRST_DATA_ROW, ws2.max_row + 1):
        raw_code = ws2.cell(row=r, column=MRP_COL_CODE).value
        if raw_code in (None, ""):
            continue
        code = sanitize_id(raw_code)
        if code not in material_ids:
            skipped += 1
            continue
        matched += 1

        for col, iso_date in incoming_dates.items():
            qty = as_number(ws2.cell(row=r, column=col).value)
            if not qty or qty <= 0:
                continue
            month_key = iso_date[:7].replace("-", "")
            purchase_orders.append(
                {
                    "id": f"MRP-{code}-{month_key}",
                    "materialId": code,
                    "poNumber": f"PO-{code}-{month_key}",
                    "orderDate": today_iso,
                    "qty": round(qty),
                    "expectedArrivalDate": iso_date,
                    "status": "outstanding",
                    "notes": "Imported from MRP sheet incoming schedule",
                }
            )

        for col, iso_date in demand_dates.items():
            qty = as_number(ws2.cell(row=r, column=col).value)
            if not qty or qty <= 0:
                continue
            month_key = iso_date[:7].replace("-", "")
            production_plan.append(
                {
                    "id": f"MRPDEM-{code}-{month_key}",
                    "materialId": code,
                    "neededByDate": iso_date,
                    "qty": round(qty),
                    "source": "MRP forecast",
                    "notes": "Imported from MRP sheet usage forecast",
                }
            )

    return {
        "materials": materials,
        "purchaseOrders": purchase_orders,
        "productionPlan": production_plan,
        "stats": {
            "materials": len(materials),
            "purchaseOrders": len(purchase_orders),
            "productionPlan": len(production_plan),
            "mrpRowsMatched": matched,
            "mrpRowsSkippedNoMaterial": skipped,
            "incomingDateColumns": incoming_dates,
            "demandDateColumns": demand_dates,
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--limit", type=int, default=None, help="Limit number of materials (for a test batch)")
    parser.add_argument("--out", default="mrp_import.json")
    args = parser.parse_args()

    result = extract(args.xlsx_path, args.limit)
    with open(args.out, "w") as f:
        json.dump(result, f, indent=2)

    print(json.dumps(result["stats"], indent=2))
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
