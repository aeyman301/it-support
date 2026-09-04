#!/usr/bin/env python3
"""
Extracts the product-level production plan from the "3. Summ FC By Harness
- PPC" sheet of PNA Technologies' monthly "MRP PLAN ALL" workbook.

Each row is a finished harness; the quantity columns are the ones marked
FIRM (the committed month) or FC (forecast months) on row 4, with their
month-end date on row 5. The same dates repeat further right for value
(quantity x selling price) columns, so the FIRM/FC marker row — not the
dates alone — is what identifies the quantity block.

This is the source plan: exploding it through the BOM reproduces the MRP
sheet's raw-material demand exactly, so entries are written without an
items array to avoid double-counting demand that already exists.

Usage:
    python3 extract_harness_plan_xlsx.py <path-to-xlsx> [--out plan.json]
"""

import argparse
import json
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SHEET = "3. Summ FC By Harness - PPC"
MARKER_ROW = 4
HEADER_ROW = 5
FIRST_DATA_ROW = 6
COL_CODE = 2
COL_NAME = 4
COL_MODEL = 5
QTY_MARKERS = {"firm", "fc"}


def sanitize_id(value) -> str:
    return str(value).strip().replace("/", "-") or "UNKNOWN"


def month_end_iso(value):
    if not isinstance(value, (datetime, date)):
        return None
    return value.strftime("%Y-%m-%d")


def find_qty_columns(ws):
    """Quantity columns carry a FIRM/FC marker on row 4 and a month-end date
    on row 5. The value columns further right repeat the dates but not the
    marker, so both are required.
    """
    cols = {}
    for c in range(1, ws.max_column + 1):
        marker = ws.cell(row=MARKER_ROW, column=c).value
        if not isinstance(marker, str) or marker.strip().lower() not in QTY_MARKERS:
            continue
        iso = month_end_iso(ws.cell(row=HEADER_ROW, column=c).value)
        if not iso:
            continue
        cols[c] = {"date": iso, "firm": marker.strip().lower() == "firm"}
    return cols


def extract(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]

    qty_cols = find_qty_columns(ws)
    if not qty_cols:
        raise ValueError("Could not find any FIRM/FC quantity columns on the harness sheet")

    entries = []
    products = {}
    firm_count = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        raw_code = ws.cell(row=r, column=COL_CODE).value
        if raw_code in (None, ""):
            continue
        code = sanitize_id(raw_code)
        name = ws.cell(row=r, column=COL_NAME).value
        name = " ".join(str(name).split()) if name not in (None, "") else code
        model = ws.cell(row=r, column=COL_MODEL).value
        # A few products are planned here but absent from the BOM sheet, so
        # emit their master data too and let the importer create any that
        # don't exist yet. No bom key: an import must never blank out a
        # recipe that's already been set.
        products.setdefault(
            code,
            {
                "id": code,
                "code": code,
                "name": name,
                "model": str(model).strip() if model not in (None, "") else "",
            },
        )

        for col, meta in qty_cols.items():
            qty = ws.cell(row=r, column=col).value
            if not isinstance(qty, (int, float)) or qty <= 0:
                continue
            month_key = meta["date"][:7].replace("-", "")
            if meta["firm"]:
                firm_count += 1
            entries.append(
                {
                    "id": f"PPC-{code}-{month_key}",
                    "productId": code,
                    "productQty": round(qty),
                    "productName": f"{code} — {name}",
                    "neededByDate": meta["date"],
                    "source": "FIRM" if meta["firm"] else "Forecast",
                    "notes": "Imported from PPC harness plan",
                }
            )

    return {
        "products": list(products.values()),
        "productionPlan": entries,
        "stats": {
            "products": len(products),
            "entries": len(entries),
            "firmEntries": firm_count,
            "forecastEntries": len(entries) - firm_count,
            "months": sorted({m["date"] for m in qty_cols.values()}),
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--out", default="harness_plan.json")
    args = parser.parse_args()

    result = extract(args.xlsx_path)
    with open(args.out, "w") as f:
        json.dump(result, f, indent=2)

    print(json.dumps(result["stats"], indent=2))
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
