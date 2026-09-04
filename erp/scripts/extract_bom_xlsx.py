#!/usr/bin/env python3
"""
Extracts the bill of materials from the "4. Summ FC By Component" sheet of
PNA Technologies' monthly "MRP PLAN ALL" workbook.

That sheet is a matrix: each row is a raw material component, each column
from column 9 onward is a finished product (harness part number), and the
cell where they meet is how much of that component one unit of that
product consumes. Trailing columns hold FIRM/FORECAST totals and per-model
summaries rather than products, so product columns are identified by
having both a description (row 16) and a model code (row 17) as text —
the totals columns carry a number and a date there instead.

Usage:
    python3 extract_bom_xlsx.py <path-to-xlsx> [--out bom_import.json]

Output shape (consumed by scripts/import_to_firestore.mjs):
    {"products": [{"id","code","name","model","bom":[{"materialId","qty"}]}]}
"""

import argparse
import json
import sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

BOM_SHEET = "4. Summ FC By Component"
PRODUCT_DESC_ROW = 16
PRODUCT_MODEL_ROW = 17
HEADER_ROW = 18
FIRST_DATA_ROW = 19
FIRST_PRODUCT_COL = 9
COL_COMPONENT_CODE = 2  # COMPONENT PART NO
COL_COMPONENT_DESC = 4  # DESCRIPTION
COL_COMPONENT_SUPPLIER = 5  # SUPPLIER
COL_COMPONENT_UOM = 8  # UOM


def sanitize_id(value) -> str:
    text = str(value).strip()
    text = text.replace("/", "-")
    return text or "UNKNOWN"


TOTALS_LABELS = {"firm", "forecast"}


def find_product_columns(ws):
    """The product columns run from column 9 up to the FIRM/FORECAST totals
    block; everything past that is totals and per-model summaries, not
    products. Stopping at that boundary rather than sniffing each column
    keeps products whose description cell holds something odd (one product
    has the number 1 where its description should be).
    """
    products = []
    for c in range(FIRST_PRODUCT_COL, ws.max_column + 1):
        code = ws.cell(row=HEADER_ROW, column=c).value
        if isinstance(code, str) and code.strip().lower() in TOTALS_LABELS:
            break  # start of the FIRM/FORECAST totals block
        if code in (None, "") or isinstance(code, (datetime, date)):
            continue
        desc = ws.cell(row=PRODUCT_DESC_ROW, column=c).value
        model = ws.cell(row=PRODUCT_MODEL_ROW, column=c).value
        name = " ".join(str(desc).split()) if isinstance(desc, str) and desc.strip() else sanitize_id(code)
        products.append(
            {
                "col": c,
                "code": sanitize_id(code),
                "name": name,
                "model": str(model).strip() if isinstance(model, str) and model.strip() else "",
            }
        )
    return products


def extract(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[BOM_SHEET]

    product_cols = find_product_columns(ws)

    # Component code per data row, collected once so the matrix is walked
    # column-by-column without re-reading the code cell for every product.
    # The row's master data comes along too: some components appear here but
    # not in the warehouse stock sheet, so the importer needs enough to
    # create them rather than leaving the BOM pointing at nothing.
    rows = []
    components = []
    seen_components = set()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        raw = ws.cell(row=r, column=COL_COMPONENT_CODE).value
        if raw in (None, ""):
            continue
        code = sanitize_id(raw)
        rows.append((r, code))
        if code in seen_components:
            continue
        seen_components.add(code)
        desc = ws.cell(row=r, column=COL_COMPONENT_DESC).value
        supplier = ws.cell(row=r, column=COL_COMPONENT_SUPPLIER).value
        uom = ws.cell(row=r, column=COL_COMPONENT_UOM).value
        supplier = str(supplier).strip() if supplier not in (None, 0, "0", "#N/A") else ""
        component = {
            "id": code,
            "code": code,
            "name": " ".join(str(desc).split()) if desc not in (None, "") else code,
            "uom": str(uom).strip() if uom not in (None, "") else "pcs",
            "leadTimeDays": 0,
            "safetyStock": 0,
            "minOrderQty": 1,
            "onHandQty": 0,
        }
        if supplier:
            component["supplier"] = supplier
        components.append(component)

    products = []
    total_lines = 0
    for p in product_cols:
        bom = []
        for r, component in rows:
            qty = ws.cell(row=r, column=p["col"]).value
            if not isinstance(qty, (int, float)):
                continue
            if qty <= 0:
                continue
            # Spreadsheet floats carry representation noise (10.616 comes
            # through as 10.616000000000001); round it off before storing.
            bom.append({"materialId": component, "qty": round(qty, 6)})
        total_lines += len(bom)
        products.append(
            {
                "id": p["code"],
                "code": p["code"],
                "name": p["name"],
                "model": p["model"],
                "bom": bom,
            }
        )

    return {
        "products": products,
        "components": components,
        "stats": {
            "productColumns": len(product_cols),
            "componentRows": len(rows),
            "distinctComponents": len(components),
            "productsWithBom": sum(1 for p in products if p["bom"]),
            "productsWithoutBom": sum(1 for p in products if not p["bom"]),
            "totalBomLines": total_lines,
        },
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--out", default="bom_import.json")
    args = parser.parse_args()

    result = extract(args.xlsx_path)
    with open(args.out, "w") as f:
        json.dump(result, f, indent=2)

    print(json.dumps(result["stats"], indent=2))
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
